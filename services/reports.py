from __future__ import annotations
from collections import defaultdict
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from db import fetch_all, now_iso

TABLES = {'inventory':'庫存報表','orders':'訂單報表','master_orders':'總單報表','shipping':'出貨報表'}

def _date_range(days=None, start='', end=''):
    if days:
        return (datetime.now()-timedelta(days=int(days))).strftime('%Y-%m-%d 00:00:00'), datetime.now().strftime('%Y-%m-%d 23:59:59')
    if start and end:
        return f'{start} 00:00:00', f'{end} 23:59:59'
    return '', ''

def query_rows(kind, days=None, start='', end=''):
    begin, finish = _date_range(days, start, end)
    params=[]
    if kind == 'shipping':
        sql = '''SELECT customer_name, material, product_text, qty, source, borrowed_from,
                        volume_total, weight_input, total_weight, operator, shipped_at AS time, note
                 FROM shipping_records'''
        if begin and finish:
            sql += ' WHERE shipped_at BETWEEN ? AND ?'; params += [begin, finish]
        sql += ' ORDER BY shipped_at DESC, id DESC'
        return fetch_all(sql, tuple(params))
    table = {'inventory':'inventory','orders':'orders','master_orders':'master_orders'}[kind]
    sql = f'''SELECT customer_name, material, product_text, qty, zone, location, operator, updated_at AS time, note
              FROM {table}'''
    if begin and finish:
        sql += ' WHERE updated_at BETWEEN ? AND ?'; params += [begin, finish]
    sql += ' ORDER BY customer_name, material, product_text, id DESC'
    return fetch_all(sql, tuple(params))

def summary(rows):
    rows=list(rows); by_customer=defaultdict(int); by_material=defaultdict(int); by_operator=defaultdict(int); total=0
    for r in rows:
        q=int(r.get('qty') or 0); total += q
        by_customer[r.get('customer_name') or '庫存'] += q
        by_material[r.get('material') or '未填材質'] += q
        by_operator[r.get('operator') or '未記錄'] += q
    top=lambda d:[{'name':k,'qty':v} for k,v in sorted(d.items(), key=lambda x:(-x[1], x[0]))[:20]]
    return {'total_rows':len(rows),'total_qty':total,'by_customer':top(by_customer),'by_material':top(by_material),'by_operator':top(by_operator)}

def workbook_for(kind, rows):
    wb=Workbook(); ws=wb.active; ws.title=TABLES.get(kind,'報表')[:31]
    ws.append([TABLES.get(kind,'報表'), f'產生時間：{now_iso()}']); ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=6); ws['A1'].font=Font(size=16,bold=True)
    if kind == 'shipping':
        headers=['客戶','材質','尺寸','件數','來源','借貨來源','材積','重量係數','總重','操作人','時間','備註']; keys=['customer_name','material','product_text','qty','source','borrowed_from','volume_total','weight_input','total_weight','operator','time','note']
    else:
        headers=['客戶','材質','尺寸','件數','A/B區','倉庫位置','操作人','時間','備註']; keys=['customer_name','material','product_text','qty','zone','location','operator','time','note']
    ws.append(headers)
    for row in rows: ws.append([row.get(k,'') for k in keys])
    fill=PatternFill('solid', fgColor='E9DFD2'); thin=Side(style='thin', color='D8CFC4')
    for c in ws[2]: c.font=Font(bold=True); c.fill=fill; c.border=Border(bottom=thin); c.alignment=Alignment(horizontal='center')
    for col in ws.columns:
        letter=get_column_letter(col[0].column); ws.column_dimensions[letter].width=min(max(10, max(len(str(c.value or '')) for c in col)+3), 42)
    ws.freeze_panes='A3'
    sm=summary(rows); ws2=wb.create_sheet('統計'); ws2.append(['統計項目','數值']); ws2.append(['總筆數',sm['total_rows']]); ws2.append(['總件數',sm['total_qty']])
    for title, items in [('客戶統計',sm['by_customer']),('材質統計',sm['by_material']),('操作人統計',sm['by_operator'])]:
        ws2.append([]); ws2.append([title,'件數'])
        for it in items: ws2.append([it['name'], it['qty']])
    ws2.column_dimensions['A'].width=28; ws2.column_dimensions['B'].width=14
    bio=BytesIO(); wb.save(bio); bio.seek(0); return bio
