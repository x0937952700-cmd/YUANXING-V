from io import BytesIO
from openpyxl import load_workbook, Workbook
from db import now_iso
from services.customers import ensure_customer
from services.items import create_item, normalize_table
from services.products import normalize_product_text, total_qty_from_text

HEADER_MAP = {
    '來源': 'table', '類型': 'table', '表': 'table',
    '客戶': 'customer_name', '客戶名稱': 'customer_name', '公司': 'customer_name',
    '材質': 'material', '商品': 'product_text', '尺寸': 'product_text', '商品資料': 'product_text',
    'A/B區': 'zone', '區域': 'zone', '倉區': 'zone', '備註': 'note'
}
TABLE_MAP = {'庫存': 'inventory', '訂單': 'orders', '總單': 'master_orders', 'inventory': 'inventory', 'orders': 'orders', 'master_orders': 'master_orders'}


def parse_import_workbook(file_bytes, default_table='inventory', operator=''):
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {'created': [], 'errors': ['檔案沒有資料']}
    headers = [str(x or '').strip() for x in rows[0]]
    keys = [HEADER_MAP.get(h, h) for h in headers]
    created, errors = [], []
    for index, raw in enumerate(rows[1:], start=2):
        data = {keys[i]: raw[i] for i in range(min(len(keys), len(raw))) if keys[i]}
        product_text = normalize_product_text(str(data.get('product_text') or '').strip())
        if not product_text:
            continue
        table_raw = str(data.get('table') or default_table or 'inventory').strip()
        table = normalize_table(TABLE_MAP.get(table_raw, table_raw))
        customer_name = str(data.get('customer_name') or '').strip()
        customer_uid = ''
        if table != 'inventory' or customer_name:
            customer = ensure_customer(customer_name or '未指定客戶')
            customer_uid, customer_name = customer['uid'], customer['name']
        try:
            for line in product_text.splitlines():
                if not line.strip():
                    continue
                item_id = create_item(table, customer_uid, customer_name, line, str(data.get('material') or '').strip(), str(data.get('zone') or '').strip(), operator, str(data.get('note') or '').strip())
                created.append({'table': table, 'id': item_id, 'customer_name': customer_name, 'product_text': line, 'qty': total_qty_from_text(line)})
        except Exception as exc:
            errors.append(f'第 {index} 列匯入失敗：{exc}')
    return {'created': created, 'errors': errors, 'created_count': len(created)}


def import_template_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = '匯入範本'
    ws.append(['來源', '客戶名稱', '材質', '商品資料', 'A/B區', '備註'])
    ws.append(['庫存', '', '白鐵', '100x30x63=115', 'A', '庫存可不填客戶'])
    ws.append(['訂單', '山益', '尤佳利', '100x30x63=220x4+223x2+44+35+221', 'B', '訂單需填客戶'])
    ws.append(['總單', '力凡', '杉木', '100x30x63=504x5+588+587+502+420+382+378+280+254+237+174', 'A', '總單需填客戶'])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
